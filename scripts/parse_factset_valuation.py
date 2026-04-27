"""
FactSet Valuation Parser
========================
Reads Valuation.xlsx:
  - Current sheet: P/E Standard, P/E Adjusted, EV/Sales, 24MF adj. EPS
  - PE_HISTORICAL sheet: 10Y monthly P/E (120 data points)
  - EV_SALES sheet: 10Y monthly EV/Sales (120 data points)

Computes:
  - Current P/E = Yahoo price / 24MF adj. EPS (column F)
  - Cross-check vs FactSet P/E Adjusted (column C) — use calculated if within 5%
  - 10Y percentile for P/E and EV/Sales
  - Sparkline data (monthly values for mini-chart)

Outputs: factset-valuation.json

Usage:
  python parse_factset_valuation.py
"""

import json
import sys
import statistics
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl --break-system-packages")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"

VALUATION_PATH = PROJECT_DIR.parent / "Files" / "NOT BACKED UP" / "FS data" / "Valuation.xlsx"
MAPPING_PATH = DATA_DIR / "ticker_mapping.json"
PRICES_PATH = DATA_DIR / "prices.json"
OUTPUT_PATH = DATA_DIR / "factset-valuation.json"


def safe_float(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "#N/A", "#N/A N/A", "#DIV/0!", "#VALUE!", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def percentile_rank(value, history):
    """What % of historical values are below this value."""
    valid = [v for v in history if v is not None]
    if not valid or value is None:
        return None
    below = sum(1 for v in valid if v < value)
    return round(below / len(valid) * 100)


def parse_historical_sheet(ws):
    """Parse a historical sheet (PE_HISTORICAL or EV_SALES) into dict of ticker -> list of monthly values."""
    # Row 1 has dates, cols 2..max_column
    dates = []
    for c in range(2, ws.max_column + 1):
        d = ws.cell(1, c).value
        if d:
            dates.append(str(d)[:10])
        else:
            dates.append(None)

    results = {}
    for r in range(2, ws.max_row + 1):
        ticker = ws.cell(r, 1).value
        if not ticker or not str(ticker).strip():
            continue
        ticker = str(ticker).strip()

        values = []
        for c in range(2, ws.max_column + 1):
            values.append(safe_float(ws.cell(r, c).value))

        results[ticker] = {"dates": dates, "values": values}

    return results


def parse_valuation():
    print(f"Reading {VALUATION_PATH}...")
    wb = openpyxl.load_workbook(str(VALUATION_PATH), data_only=True)

    # --- Current sheet ---
    ws_cur = wb["Current"]
    current_data = {}
    for r in range(2, ws_cur.max_row + 1):
        ticker = ws_cur.cell(r, 1).value
        if not ticker or not str(ticker).strip():
            continue
        ticker = str(ticker).strip()

        pe_standard = safe_float(ws_cur.cell(r, 2).value)
        pe_adjusted = safe_float(ws_cur.cell(r, 3).value)
        ev_sales = safe_float(ws_cur.cell(r, 4).value)
        eps_24mf = safe_float(ws_cur.cell(r, 6).value)

        current_data[ticker] = {
            "pe_standard": pe_standard,
            "pe_adjusted": pe_adjusted,
            "ev_sales_current": ev_sales,
            "eps_24mf": eps_24mf,
        }

    # --- Historical sheets ---
    print("  Parsing PE_HISTORICAL...")
    pe_hist = parse_historical_sheet(wb["PE_HISTORICAL"])
    print("  Parsing EV_SALES...")
    ev_sales_hist = parse_historical_sheet(wb["EV_SALES"])

    wb.close()
    return current_data, pe_hist, ev_sales_hist


def build_output(current_data, pe_hist, ev_sales_hist, prices):
    """Combine current + historical + Yahoo prices into output."""
    with open(str(MAPPING_PATH)) as f:
        mapping = json.load(f)

    output = {
        "_meta": {
            "source": "Valuation.xlsx",
            "description": "Valuation multiples: current P/E, EV/Sales, 10Y history, percentiles",
        }
    }

    matched = 0
    for universe_ticker, m in mapping["stocks"].items():
        fs_ticker = m["fs"]

        cur = current_data.get(fs_ticker, {})
        pe_h = pe_hist.get(fs_ticker, {})
        ev_h = ev_sales_hist.get(fs_ticker, {})

        if not cur:
            print(f"  WARN: {universe_ticker} (fs={fs_ticker}) not in Current sheet")
            continue

        # Get Yahoo price
        yahoo_price = None
        if universe_ticker in prices:
            yahoo_price = prices[universe_ticker].get("price")

        # Calculate P/E = Yahoo price / 24MF adj. EPS
        eps = cur.get("eps_24mf")
        pe_calculated = None
        if yahoo_price and eps and abs(eps) > 0.001:
            pe_calculated = yahoo_price / eps

        # Cross-check vs FactSet P/E Adjusted (column C)
        pe_fs = cur.get("pe_adjusted")
        pe_final = None
        pe_source = None

        if pe_calculated is not None and pe_fs is not None:
            # Within 5% tolerance — use calculated
            if abs(pe_fs) > 0.01:
                diff_pct = abs(pe_calculated - pe_fs) / abs(pe_fs) * 100
                if diff_pct <= 5:
                    pe_final = round(pe_calculated, 1)
                    pe_source = "calculated"
                else:
                    # Use judgement: P/E normally 8-30x
                    if 5 <= pe_calculated <= 80 and 5 <= pe_fs <= 80:
                        # Both reasonable — prefer calculated (uses latest price)
                        pe_final = round(pe_calculated, 1)
                        pe_source = "calculated_divergent"
                    elif 5 <= pe_calculated <= 80:
                        pe_final = round(pe_calculated, 1)
                        pe_source = "calculated_only_valid"
                    elif 5 <= pe_fs <= 80:
                        pe_final = round(pe_fs, 1)
                        pe_source = "factset_only_valid"
                    else:
                        pe_final = round(pe_calculated, 1)
                        pe_source = "calculated_both_outlier"
            else:
                pe_final = round(pe_calculated, 1)
                pe_source = "calculated_fs_zero"
        elif pe_calculated is not None:
            pe_final = round(pe_calculated, 1)
            pe_source = "calculated_no_fs"
        elif pe_fs is not None:
            pe_final = round(pe_fs, 1)
            pe_source = "factset_no_calc"

        # Historical P/E sparkline + percentile
        pe_history = pe_h.get("values", [])
        pe_valid_history = [v for v in pe_history if v is not None and 0 < v < 200]
        pe_percentile = percentile_rank(pe_final, pe_valid_history) if pe_final else None

        # Sparkline: last 120 months, most recent first
        pe_sparkline = [round(v, 1) if v is not None and 0 < v < 200 else None for v in pe_history]

        # EV/Sales sparkline + percentile
        ev_history = ev_h.get("values", [])
        ev_valid_history = [v for v in ev_history if v is not None and v > 0]
        ev_current = cur.get("ev_sales_current")
        ev_percentile = percentile_rank(ev_current, ev_valid_history) if ev_current else None
        ev_sparkline = [round(v, 2) if v is not None and v > 0 else None for v in ev_history]

        # 10Y range
        pe_10y_low = round(min(pe_valid_history), 1) if pe_valid_history else None
        pe_10y_high = round(max(pe_valid_history), 1) if pe_valid_history else None
        ev_10y_low = round(min(ev_valid_history), 2) if ev_valid_history else None
        ev_10y_high = round(max(ev_valid_history), 2) if ev_valid_history else None

        output[universe_ticker] = {
            "pe_current": pe_final,
            "pe_source": pe_source,
            "pe_fs_adjusted": round(pe_fs, 1) if pe_fs is not None else None,
            "pe_calculated": round(pe_calculated, 1) if pe_calculated is not None else None,
            "eps_24mf": round(eps, 4) if eps is not None else None,
            "ev_sales_current": round(ev_current, 2) if ev_current is not None else None,
            "pe_percentile": pe_percentile,
            "ev_sales_percentile": ev_percentile,
            "pe_10y_low": pe_10y_low,
            "pe_10y_high": pe_10y_high,
            "ev_sales_10y_low": ev_10y_low,
            "ev_sales_10y_high": ev_10y_high,
            "pe_sparkline": pe_sparkline,
            "ev_sales_sparkline": ev_sparkline,
        }
        matched += 1

    print(f"  Matched {matched}/{len(mapping['stocks'])} universe stocks")
    return output


def main():
    current_data, pe_hist, ev_sales_hist = parse_valuation()
    print(f"  Parsed {len(current_data)} stocks from Current, {len(pe_hist)} from PE_HISTORICAL, {len(ev_sales_hist)} from EV_SALES")

    # Load prices for P/E calculation
    prices = {}
    if PRICES_PATH.exists():
        with open(str(PRICES_PATH)) as f:
            prices = json.load(f)
        print(f"  Loaded {len(prices)} stocks from prices.json")
    else:
        print("  WARN: prices.json not found — P/E calculation will use FactSet only")

    output = build_output(current_data, pe_hist, ev_sales_hist, prices)

    with open(str(OUTPUT_PATH), "w") as f:
        json.dump(output, f, indent=2)
    print(f"  Written: {OUTPUT_PATH} ({OUTPUT_PATH.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
