"""
FactSet SSEM Parser
===================
Reads SSEM.xlsx, computes revision % changes for EPS/EBITDA/Sales/TP/Buy
over L1M, L3M, L6M, L12M. Computes momentum count. Outputs factset-ssem.json.

Column layout in SSEM.xlsx 'FS' sheet:
  A: Ticker
  B-F: 2027 EPS at 5 dates (current, -1M, -3M, -6M, -12M)
  G: blank
  H-L: 2027 EBITDA at 5 dates
  M: blank
  N-R: 2027 Sales at 5 dates (NB: col R header says EBITDA but is Sales -12M)
  S: blank
  T-W: TP (current, -1M, -3M, -6M)
  X: blank
  Y-AA: % Buy/Hold/Sell current
  AB: blank
  AC-AD: % Buy -1M, -3M

Usage:
  python parse_factset_ssem.py
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl --break-system-packages")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"

# Paths
SSEM_PATH = PROJECT_DIR.parent / "Files" / "NOT BACKED UP" / "FS data" / "SSEM.xlsx"
MAPPING_PATH = DATA_DIR / "ticker_mapping.json"
OUTPUT_PATH = DATA_DIR / "factset-ssem.json"


def safe_float(v):
    """Convert cell value to float, return None if not numeric."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "#N/A", "#N/A N/A", "#DIV/0!", "#VALUE!", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_ssem_from_export(path):
    """Parse the pre-exported JSON from export_ssem_local.py."""
    with open(str(path)) as f:
        data = json.load(f)
    results = {}
    for row in data["rows"]:
        ticker = row["ticker"]
        eps = row["eps"]
        ebitda = row["ebitda"]
        sales = row["sales"]
        tp = row["tp"]

        eps_rev = {
            "L1M": pct_change(eps[0], eps[1]) if len(eps) > 1 else None,
            "L3M": pct_change(eps[0], eps[2]) if len(eps) > 2 else None,
            "L6M": pct_change(eps[0], eps[3]) if len(eps) > 3 else None,
            "L12M": pct_change(eps[0], eps[4]) if len(eps) > 4 else None,
        }
        ebitda_rev = {
            "L1M": pct_change(ebitda[0], ebitda[1]) if len(ebitda) > 1 else None,
            "L3M": pct_change(ebitda[0], ebitda[2]) if len(ebitda) > 2 else None,
            "L6M": pct_change(ebitda[0], ebitda[3]) if len(ebitda) > 3 else None,
            "L12M": pct_change(ebitda[0], ebitda[4]) if len(ebitda) > 4 else None,
        }
        sales_rev = {
            "L1M": pct_change(sales[0], sales[1]) if len(sales) > 1 else None,
            "L3M": pct_change(sales[0], sales[2]) if len(sales) > 2 else None,
            "L6M": pct_change(sales[0], sales[3]) if len(sales) > 3 else None,
            "L12M": pct_change(sales[0], sales[4]) if len(sales) > 4 else None,
        }
        tp_rev = {
            "L1M": pct_change(tp[0], tp[1]) if len(tp) > 1 else None,
            "L3M": pct_change(tp[0], tp[2]) if len(tp) > 2 else None,
            "L6M": pct_change(tp[0], tp[3]) if len(tp) > 3 else None,
            "L12M": None,
        }

        buy_current = row.get("buy_current")
        buy_1m = row.get("buy_1m")
        buy_3m = row.get("buy_3m")
        buy_6m = row.get("buy_6m")
        buy_rev = {
            "L1M": round(buy_current - buy_1m, 1) if buy_current is not None and buy_1m is not None else None,
            "L3M": round(buy_current - buy_3m, 1) if buy_current is not None and buy_3m is not None else None,
            "L6M": round(buy_current - buy_6m, 1) if buy_current is not None and buy_6m is not None else None,
            "L12M": None,
        }

        momentum = 0
        for tf in ("L1M", "L3M", "L6M"):
            for rev_dict in (eps_rev, ebitda_rev, sales_rev, tp_rev, buy_rev):
                v = rev_dict.get(tf)
                if v is not None:
                    if v > 0:
                        momentum += 1
                    elif v < 0:
                        momentum -= 1

        def round_rev(d):
            return {k: (round(v) if v is not None else None) for k, v in d.items()}

        results[ticker] = {
            "eps_rev": round_rev(eps_rev),
            "ebitda_rev": round_rev(ebitda_rev),
            "sales_rev": round_rev(sales_rev),
            "tp_rev": round_rev(tp_rev),
            "buy_rev": round_rev(buy_rev),
            "buy_pct": round(buy_current) if buy_current is not None else None,
            "momentum": momentum,
            "raw": {
                "eps_current": eps[0] if eps else None,
                "ebitda_current": ebitda[0] if ebitda else None,
                "sales_current": sales[0] if sales else None,
                "tp_current": tp[0] if tp else None,
            },
        }
    print(f"  Parsed {len(results)} stocks from export")
    return results


def pct_change(current, prior):
    """Calculate % change: (current - prior) / abs(prior) * 100. Returns None if either is None or prior is 0."""
    if current is None or prior is None:
        return None
    if abs(prior) < 1e-10:
        return None
    return (current - prior) / abs(prior) * 100.0


def parse_ssem():
    """Parse SSEM data. Tries exported JSON first, falls back to xlsx."""
    # Try the pre-exported JSON first (avoids FactSet zip format issues)
    export_path = DATA_DIR / "ssem-raw-export.json"
    if export_path.exists():
        print(f"Reading pre-exported {export_path}...")
        return parse_ssem_from_export(export_path)

    print(f"Reading {SSEM_PATH}...")
    wb = openpyxl.load_workbook(str(SSEM_PATH), data_only=True)
    ws = wb["FS"]

    results = {}

    for r in range(2, ws.max_row + 1):
        ticker = ws.cell(r, 1).value
        if not ticker or not str(ticker).strip():
            continue
        ticker = str(ticker).strip()

        # EPS: cols 2-6 (current, -1M, -3M, -6M, -12M)
        eps = [safe_float(ws.cell(r, c).value) for c in range(2, 7)]

        # EBITDA: cols 8-12
        ebitda = [safe_float(ws.cell(r, c).value) for c in range(8, 13)]

        # Sales: cols 14-18
        sales = [safe_float(ws.cell(r, c).value) for c in range(14, 19)]

        # TP: cols 20-23 (current, -1M, -3M, -6M) — no -12M for TP
        tp = [safe_float(ws.cell(r, c).value) for c in range(20, 24)]

        # % Buy: col 25 = current, col 29 = -1M, col 30 = -3M, col 31 = -6M (added by Richard 23-Apr-26)
        buy_current = safe_float(ws.cell(r, 25).value)
        buy_1m = safe_float(ws.cell(r, 29).value)
        buy_3m = safe_float(ws.cell(r, 30).value)
        buy_6m = safe_float(ws.cell(r, 31).value)

        # Compute revision %s
        # EPS revisions
        eps_rev = {
            "L1M": pct_change(eps[0], eps[1]),
            "L3M": pct_change(eps[0], eps[2]),
            "L6M": pct_change(eps[0], eps[3]),
            "L12M": pct_change(eps[0], eps[4]),
        }

        # EBITDA revisions
        ebitda_rev = {
            "L1M": pct_change(ebitda[0], ebitda[1]),
            "L3M": pct_change(ebitda[0], ebitda[2]),
            "L6M": pct_change(ebitda[0], ebitda[3]),
            "L12M": pct_change(ebitda[0], ebitda[4]),
        }

        # Sales revisions
        sales_rev = {
            "L1M": pct_change(sales[0], sales[1]),
            "L3M": pct_change(sales[0], sales[2]),
            "L6M": pct_change(sales[0], sales[3]),
            "L12M": pct_change(sales[0], sales[4]),
        }

        # TP revisions (no L12M available)
        tp_rev = {
            "L1M": pct_change(tp[0], tp[1]),
            "L3M": pct_change(tp[0], tp[2]),
            "L6M": pct_change(tp[0], tp[3]),
            "L12M": None,
        }

        # Buy % change (absolute change in percentage points)
        buy_rev = {
            "L1M": round(buy_current - buy_1m, 1) if buy_current is not None and buy_1m is not None else None,
            "L3M": round(buy_current - buy_3m, 1) if buy_current is not None and buy_3m is not None else None,
            "L6M": round(buy_current - buy_6m, 1) if buy_current is not None and buy_6m is not None else None,
            "L12M": None,
        }

        # Momentum count: for L1M, L3M, L6M (not L12M),
        # any revision > 0 gets +1, < 0 gets -1
        momentum = 0
        for tf in ("L1M", "L3M", "L6M"):
            for rev_dict in (eps_rev, ebitda_rev, sales_rev, tp_rev, buy_rev):
                v = rev_dict.get(tf)
                if v is not None:
                    if v > 0:
                        momentum += 1
                    elif v < 0:
                        momentum -= 1

        # Round all revisions to integers (0dp)
        def round_rev(d):
            return {k: (round(v) if v is not None else None) for k, v in d.items()}

        results[ticker] = {
            "eps_rev": round_rev(eps_rev),
            "ebitda_rev": round_rev(ebitda_rev),
            "sales_rev": round_rev(sales_rev),
            "tp_rev": round_rev(tp_rev),
            "buy_rev": round_rev(buy_rev),
            "buy_pct": round(buy_current) if buy_current is not None else None,
            "momentum": momentum,
            "raw": {
                "eps_current": eps[0],
                "ebitda_current": ebitda[0],
                "sales_current": sales[0],
                "tp_current": tp[0],
            },
        }

    wb.close()
    return results


def filter_to_universe(all_ssem):
    """Filter SSEM data to our universe stocks.

    Session 11 (01-May-26): Switched from ticker_mapping.json (29 hand-built entries)
    to direct match against universe.json (976 stocks). SSEM.xlsx uses the same
    FactSet ticker format as universe.json, so ~96% match directly. Falls back to
    ticker_mapping.json for the few exceptions (e.g. CARL.B-DK in SSEM vs CARLB-DK
    in universe).
    """
    universe_path = DATA_DIR / "universe.json"
    with open(str(universe_path)) as f:
        universe = json.load(f)
    universe_list = universe if isinstance(universe, list) else universe.get("stocks", [])

    # Optional fallback mapping for ticker-format mismatches (loaded only if exists)
    fallback_mapping = {}
    if MAPPING_PATH.exists():
        with open(str(MAPPING_PATH)) as f:
            mp = json.load(f)
        for univ_t, m in mp.get("stocks", {}).items():
            fs_t = m.get("fs")
            if fs_t and fs_t != univ_t:
                fallback_mapping[univ_t] = fs_t

    output = {"_meta": {"source": "SSEM.xlsx", "description": "SS earnings momentum revisions"}}

    # Dual-class share suffix candidates for FactSet ticker convention.
    # Universe uses simplified form (e.g. ASSA-SE); SSEM uses .A-SE / .B-SE / .C-SE for Nordic + UK dual-class shares.
    # Only applies to country codes that have known dual-class conventions; cross-country fallback (e.g. -SE -> -DK)
    # is intentionally NOT done because it risks silently matching different companies on similar tickers.
    DUAL_CLASS_COUNTRIES = {"-SE", "-DK", "-NO", "-FI", "-GB"}

    def dual_class_candidates(ticker):
        """Yield SSEM-format candidates for a universe ticker."""
        for suffix in DUAL_CLASS_COUNTRIES:
            if ticker.endswith(suffix):
                root = ticker[: -len(suffix)]
                # Try .B first (more common), then .A, then .C
                yield f"{root}.B{suffix}"
                yield f"{root}.A{suffix}"
                yield f"{root}.C{suffix}"
                break

    matched_direct = 0
    matched_dual_class = 0
    matched_fallback = 0
    missing = []
    dual_class_resolutions = []
    for s in universe_list:
        univ_ticker = s.get("ticker")
        if not univ_ticker:
            continue
        # Try direct match first
        if univ_ticker in all_ssem:
            output[univ_ticker] = all_ssem[univ_ticker]
            matched_direct += 1
            continue
        # Try dual-class suffix variants (.A / .B / .C)
        resolved = None
        for cand in dual_class_candidates(univ_ticker):
            if cand in all_ssem:
                resolved = cand
                break
        if resolved is not None:
            output[univ_ticker] = all_ssem[resolved]
            matched_dual_class += 1
            dual_class_resolutions.append((univ_ticker, resolved))
            continue
        # Try fallback mapping (legacy ticker_mapping.json)
        if univ_ticker in fallback_mapping:
            fs_t = fallback_mapping[univ_ticker]
            if fs_t in all_ssem:
                output[univ_ticker] = all_ssem[fs_t]
                matched_fallback += 1
                continue
        missing.append(univ_ticker)

    total_matched = matched_direct + matched_dual_class + matched_fallback
    print(f"  Direct matches: {matched_direct}")
    print(f"  Dual-class suffix matches: {matched_dual_class}")
    if dual_class_resolutions:
        print(f"    First 5 dual-class: {dual_class_resolutions[:5]}")
    print(f"  Fallback-mapping matches: {matched_fallback}")
    print(f"  Missing (no SSEM data): {len(missing)}")
    if missing:
        print(f"    All missing: {missing}")
    print(f"  Total matched: {total_matched}/{len(universe_list)} universe stocks ({100*total_matched/len(universe_list):.1f}%)")
    return output


def main():
    all_ssem = parse_ssem()
    print(f"  Parsed {len(all_ssem)} stocks from SSEM.xlsx")

    filtered = filter_to_universe(all_ssem)

    with open(str(OUTPUT_PATH), "w") as f:
        json.dump(filtered, f, indent=2)
    print(f"  Written: {OUTPUT_PATH} ({OUTPUT_PATH.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
