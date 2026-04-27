r"""
SSEM Local Exporter - Run on Richard's machine
================================================
Reads SSEM.xlsx via openpyxl and writes a clean JSON that
the sandbox can always read (no FactSet zip format issues).

Run this on your machine whenever you refresh SSEM.xlsx:
  cd C:\Users\richb\Documents\COWORK\master-dashboard
  python scripts\export_ssem_local.py

Then the sandbox parse_factset_ssem.py will read the exported JSON instead.
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Run: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"

SSEM_PATH = PROJECT_DIR.parent / "Files" / "NOT BACKED UP" / "FS data" / "SSEM.xlsx"
OUTPUT_PATH = DATA_DIR / "ssem-raw-export.json"


def safe_float(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "#N/A", "#N/A N/A", "#DIV/0!", "#VALUE!", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def main():
    print(f"Reading {SSEM_PATH}...")
    wb = openpyxl.load_workbook(str(SSEM_PATH), data_only=True)
    ws = wb["FS"]

    rows = []
    for r in range(2, ws.max_row + 1):
        ticker = ws.cell(r, 1).value
        if not ticker or not str(ticker).strip():
            continue
        row = {"ticker": str(ticker).strip()}
        # EPS: cols 2-6
        row["eps"] = [safe_float(ws.cell(r, c).value) for c in range(2, 7)]
        # EBITDA: cols 8-12
        row["ebitda"] = [safe_float(ws.cell(r, c).value) for c in range(8, 13)]
        # Sales: cols 14-18
        row["sales"] = [safe_float(ws.cell(r, c).value) for c in range(14, 19)]
        # TP: cols 20-23
        row["tp"] = [safe_float(ws.cell(r, c).value) for c in range(20, 24)]
        # Buy: col 25=current, 29=-1M, 30=-3M, 31=-6M
        row["buy_current"] = safe_float(ws.cell(r, 25).value)
        row["buy_1m"] = safe_float(ws.cell(r, 29).value)
        row["buy_3m"] = safe_float(ws.cell(r, 30).value)
        row["buy_6m"] = safe_float(ws.cell(r, 31).value)
        rows.append(row)

    wb.close()

    output = {"_meta": {"source": str(SSEM_PATH), "count": len(rows)}, "rows": rows}
    with open(str(OUTPUT_PATH), "w") as f:
        json.dump(output, f)
    print(f"Exported {len(rows)} stocks to {OUTPUT_PATH} ({OUTPUT_PATH.stat().st_size:,} bytes)")
    print("Now run: python scripts/parse_factset_ssem.py")


if __name__ == "__main__":
    main()
